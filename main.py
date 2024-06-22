from openpyxl import Workbook, load_workbook
from browser import StockBrowser
from handler import ExcelHandler

def main():
    stocks = [
        "ΚΑΡΕΛ",
        "ΜΕΡΚΟ",
        "ΑΝΔΡΟ",
        "ΓΕΚΤΕΡΝΑ",
        "ΠΑΠ",
        "ΜΙΓ",
        "ΔΟΜΙΚ",
        "ΙΝΚΑΤ",
        "ΑΤΤΙΚΑ",
        "ΚΛΜ",
    ]
    headers = {
        "Close": None,
        "Performance": None,
        "Performance %": None,
        "Market Open": None,
        "High": None,
        "Low": None,
        "Volume": None,
        "Value": None,
        "Actions": None,
    }
    filename = "StockData.xlsx"

    try:
        wb = load_workbook(filename)
        if "StocksData" in wb.sheetnames:
            wb.remove(wb["StocksData"])
    except Exception:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    wb.create_sheet("StocksData")
    ws = wb["StocksData"]
    ws.sheet_view.showGridLines = False

    browser = StockBrowser(stocks, headers)
    handler = ExcelHandler(wb, ws, headers, stocks, browser)
    handler.save_changes(filename)

if __name__ == "__main__":
    main()
