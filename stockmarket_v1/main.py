from openpyxl import load_workbook

from stockbrowser import StockBrowser
from excelhandler import ExcelHandler


def main():
    stocks = [
        "ΚΑΡΕΛ",
        "ΜΕΡΚΟ",
        "ΑΝΔΡΟ",
        "ΓΕΚΤΕΡΝΑ",
        "ΠΑΠ",
        "ΜΙΓ",
        "ΔΟΜΙΚ",
        "ΙΝΚΑΤ",
        "ΑΤΤΙΚΑ",
        "ΚΛΜ",
    ]
    headers = {
        "Close": None,
        "Performance": None,
        "Performance %": None,
        "Market Open": None,
        "High": None,
        "Low": None,
        "Volume": None,
        "Value": None,
        "Actions": None,
    }
    filename = "StockData.xlsx"

    wb = load_workbook(filename)
    wb.remove(wb["StocksData"])
    wb.create_sheet("StocksData")
    ws = wb["StocksData"]
    ws.sheet_view.showGridLines = False
    
    browser = StockBrowser(stocks, headers)
    handler = ExcelHandler(wb, ws, headers, stocks, browser)
    handler.style_cells()
    handler.align_cells()
    handler.modify_cell_font()
    handler.fix_notation()
    handler.create_history_worksheets()
    handler.fill_history()
    handler.style_history()
    handler.convert_to_value()
    handler.save_changes(filename)


if __name__ == "__main__":
    main()
