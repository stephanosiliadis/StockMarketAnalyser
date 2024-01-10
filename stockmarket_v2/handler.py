from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelHandler(object):
    def __init__(self, filename, downloader):
        self.filename = filename
        self.downloader = downloader
        self.data = self.downloader.download_data()
        self.downloader.get_xlsx(filename)
        self.wb = load_workbook(self.filename)
        self.active_ws = self.wb.active
        self.active_ws.title = "StockData"
        self.total_rows = 3 + len(self.data)
        self.total_cols = 1 + 6 * len(self.downloader.tickers)

    def style_cells(self):
        # Get each cell and change its bg color and font
        self.active_ws.sheet_view.showGridLines = False
        for row in range(1, self.total_rows + 1):
            for col in range(1, self.total_cols + 1):
                char = get_column_letter(col)
                self.active_ws.column_dimensions[char].width = 15 if char != "A" else 23
                cell = f"{char}{row}"
                self.active_ws[cell].font = Font(color="FFFFFFFF", size=14)
                self.active_ws[cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                self.active_ws[cell].fill = PatternFill(
                    start_color="FF333333", end_color="FF333333", fill_type="solid"
                )

    def align_cells(self):
        for row in range(4, self.total_rows + 1):
            for col in range(2, self.total_cols + 1):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                self.active_ws[cell].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

    def round_values(self):
        for row in range(4, self.total_rows + 1):
            for col in range(2, self.total_cols + 1):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                val = round(self.active_ws[cell].value, 2)
                self.active_ws[cell] = val

    def save_changes(self):
        self.wb.save(self.filename)
        print("Changes saved!")
