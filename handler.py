from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd


class ExcelHandler(object):
    def __init__(self, wb, ws, headers, stocks, browser):
        self.wb = wb
        self.ws = ws
        self.headers = headers
        self.stocks = stocks
        self.stock_data = browser.get_stock_data()
        self.answer = input(
            "Would you like to get the full history of your stocks? (y/n): "
        )
        if self.answer == "y":
            self.stock_history = browser.get_stock_history()
        else:
            self.stock_history = browser.get_stock_latest_history()
        print(pd.DataFrame.from_dict(self.stock_history))

        # Initialize the excel workbook and fill in the data
        self.ws.append(["Symbols"] + list(headers.keys()))
        for col in range(len(stocks)):
            self.ws.append([stocks[col]] + list(self.stock_data[stocks[col]].values()))

        # Modify the width of all columns
        for col in range(1, len(headers) + 2):
            self.ws.column_dimensions[get_column_letter(col)].width = 17

        self.style_cells()
        self.align_cells()
        self.modify_cell_font()
        self.fix_notation()
        self.create_history_worksheets()
        self.fill_history()
        self.style_history()

    def style_cells(self):
        # Make cell bg black and font white and align center
        for row in range(1, len(self.stocks) + 2):
            for col in range(1, len(self.headers) + 2):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                self.ws[cell].font = Font(color="FFFFFFFF", size=14)
                self.ws[cell].fill = PatternFill(
                    start_color="FF333333", end_color="FF333333", fill_type="solid"
                )
                self.ws[cell].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    def align_cells(self):
        # Performance symbol alignment:
        for row in range(2, len(self.stocks) + 2):
            cell = f"A{row}"
            self.ws[cell].alignment = Alignment(horizontal="left", vertical="center")

        # Performance numeric alignments:
        for row in range(2, len(self.stocks) + 2):
            for col in range(2, len(self.headers) + 2):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                self.ws[cell].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

    def modify_cell_font(self):
        # Make headings bold:
        for col in range(1, len(self.headers) + 2):
            char = get_column_letter(col)
            cell = f"{char}1"
            self.ws[cell].font = Font(color="FFFFFFFF", size=14, bold=True)

        # Make cell font color represent performance +/-
        for row in range(2, len(self.stocks) + 2):
            for col in range(3, 5):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                val = self.ws[cell].value
                try:
                    val = val.replace(",", ".")
                    val = val.replace("%", "")
                except Exception:
                    pass
                if float(val) > 0:
                    color = "FF00FF00"
                elif float(val) < 0:
                    color = "FFFF0000"
                else:
                    color = "FFFFFFFF"
                self.ws[cell].font = Font(color=color, size=14)

    def fix_notation(self):
        # Convert back to Greek notation for float values
        for row in range(2, len(self.stocks) + 2):
            for col in range(2, 5):
                char = get_column_letter(col)
                cell = f"{char}{row}"
                val = self.ws[cell].value
                try:
                    val = val.replace(".", ",")
                except Exception:
                    pass

                self.ws[cell] = val

    def create_history_worksheets(self):
        for stock in self.stocks:
            sheet_name = f"{stock} History"
            if sheet_name not in self.wb.get_sheet_names():
                self.wb.create_sheet(sheet_name)
            else:
                pass

    # TODO: Reverse the history table order so that the latest history is shown last
    def fill_history(self):
        ws_names = self.wb.get_sheet_names()[1:]
        for name in ws_names:
            if name != "StocksData":
                ws = self.wb[name]
            else:
                continue
            stock = name.replace(" History", "")
            data = self.stock_history[stock]
            history_headers = [
                "Close",
                "% Performance",
                "Market Open",
                "High",
                "Low",
                "Volume",
                "Value",
            ]
            # Put the history headers in the 1st row
            ws["A1"] = "Date"
            for col, header in enumerate(history_headers):
                char = get_column_letter(col + 2)
                cell = f"{char}1"
                ws.column_dimensions[char].width = 17
                ws[cell] = header

            # Fill in the history
            for row, date in enumerate(data.keys()):
                if self.answer != "y" and date != ws["A2"].value:
                    ws.insert_rows(1)
                    cell = f"A{row+2}"
                    ws[cell] = date
                    stock_data = data[date]
                    for col in range(2, len(stock_data) + 2):
                        char = get_column_letter(col)
                        cell = f"{char}{row+2}"
                        ws[cell] = stock_data[list(stock_data.keys())[col - 2]]
                else:
                    cell = f"A{row+2}"
                    ws[cell] = date
                    stock_data = data[date]
                    for col in range(2, len(stock_data) + 2):
                        char = get_column_letter(col)
                        cell = f"{char}{row+2}"
                        ws[cell] = stock_data[list(stock_data.keys())[col - 2]]

    def style_history(self):
        ws_names = self.wb.get_sheet_names()[1:]
        for name in ws_names:
            if name != "StocksData":
                ws = self.wb[name]
            else:
                continue
            # Make the column width to be 10 and hide the Grid Lines
            ws.column_dimensions["A"].width = 15
            ws.sheet_view.showGridLines = False
            # Apply colors, basic fonts and basic alignment
            stock = name.replace(" History", "")
            data = self.stock_history[stock]
            history_headers = [
                "Close",
                "% Performance",
                "Market Open",
                "High",
                "Low",
                "Volume",
                "Value",
            ]
            for row in range(1, len(data.keys()) + 2):
                for col in range(1, len(history_headers) + 2):
                    char = get_column_letter(col)
                    cell = f"{char}{row}"
                    ws[cell].font = Font(color="FFFFFFFF", size=14)
                    ws[cell].fill = PatternFill(
                        start_color="FF333333", end_color="FF333333", fill_type="solid"
                    )
                    ws[cell].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

            # Make headings bold
            for col in range(1, len(history_headers) + 2):
                char = get_column_letter(col)
                cell = f"{char}1"
                ws[cell].font = Font(color="FFFFFFFF", size=14, bold=True)

            # Align cells properly
            for row in range(2, len(data.keys()) + 2):
                for col in range(2, len(history_headers) + 2):
                    char = get_column_letter(col)
                    cell = f"{char}{row}"
                    ws[cell].alignment = Alignment(
                        horizontal="right", vertical="center"
                    )

            # Change font color for the % Performance column (C)
            for row in range(2, len(data.keys()) + 2):
                cell = f"C{row}"
                percentage_performance = (
                    ws[cell].value.replace(",", ".").replace("%", "")
                )
                if float(percentage_performance) > 0:
                    color = "FF00FF00"
                elif float(percentage_performance) < 0:
                    color = "FFFF0000"
                else:
                    color = "FFFFFFFF"
                ws[cell].font = Font(color=color, size=14)

    # TODO: Fill func
    def convert_to_value(self):
        pass

    # TODO: Fill func
    def macd_indicator(self):
        """
        MACD = (12-day-avg) - (16-day-avg)
        """
        pass

    def save_changes(self, path):
        self.wb.save(path)
        print("Workbook is ready!")
